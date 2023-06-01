from openpyxl.reader.excel import load_workbook

from src.domain.customer.Customer import Customer
from src.domain.customer.Customers import Customers
from src.domain.menu.Menu import Menu
from src.util.Validator import Validator
import openpyxl


class CustomerMenu(Menu):
    _instance = None
    __customer_service: Customers
    __prev_add_list: list[Customer]

    def __new__(cls, *args, **kwargs):
        if CustomerMenu._instance is None:
            CustomerMenu._instance = super().__new__(cls)
            return CustomerMenu._instance

    def __init__(self, customer_service: Customers):
        self.__customer_service = customer_service
        self.__prev_add_list = list()

    @staticmethod
    def get_instance():
        if CustomerMenu._instance is None:
            CustomerMenu._instance = CustomerMenu(Customers.get_instance())
            return CustomerMenu._instance
        else:
            return CustomerMenu._instance

    def print_menu(self):
        mode = True
        while mode:
            print("고객 관리 메뉴")
            print("1. 고객 추가")
            print("2. 고객 삭제")
            print("3. 고객 수정")
            print("4. 고객 조회")
            print("5. 돌아가기")
            print("0. 종료")
            menu_input = self.get_input()
            if menu_input == 5:
                mode = False
            self.select_menu(menu_input)

    def select_menu(self, menu_input: int):
        if menu_input == 1:
            self.add_customer_select_menu()
        elif menu_input == 2:
            self.delete_customer_select()
        elif menu_input == 3:
            self.update_customer()
        elif menu_input == 4:
            self.print_customers()

    def add_customer_select_menu(self):
        while True:
            print("고객을 추가합니다.")
            print("1. 직접 입력")
            print("2. 엑셀 파일로 추가")
            print("0. 돌아가기")
            try:
                menu_input = int(input())
                if Validator.validate_is_go_back_menu(menu_input):
                    return
                if menu_input == 1:
                    self.add_customer()
                elif menu_input == 2:
                    self.add_customer_by_xl()
                else:
                    print("잘못된 입력입니다. 다시 입력해주세요.")
            except ValueError:
                print("숫자만 입력해주세요.")

    def add_customer(self):
        print("고객 명단을 추가합니다.")
        print("추가를 원하는 인원 수를 입력해주세요. 돌아가기를 원할 시 0을 입력해주세요.")
        try:
            num = int(input())
            if Validator.validate_is_go_back_menu(num):
                return
        except ValueError:
            print("숫자만 입력해주세요. 돌아갑니다.")
            return
        self.add_customer_recall(num)

    def add_customer_by_xl(self):
        print("엑셀 파일로 고객 명단을 추가합니다.")
        print("고객 아이디, 고객 이름, 고객 이용 금액과 시간의 엑셀 파일을 선택해주세요.")
        print("돌아가기를 원할 시 0을 입력해주세요.")
        row_list = list()
        while True:
            file_path = input("파일 경로를 입력해주세요.")
            if Validator.validate_is_go_back_menu(file_path):
                return
            if Validator.validate_file_path(file_path):
                wb = load_workbook(file_path)
                break
            else:
                print("파일 경로가 잘못되었습니다.")
        print(f"파일을 읽어왔습니다.")
        while True:
            sheet_name = input(f"시트 이름을 입력해주세요." + "\n" + f"시트 목록 : {wb.sheetnames}")
            if Validator.validate_is_go_back_menu(sheet_name):
                return
            if Validator.validate_sheet_name(wb, sheet_name):
                break
            else:
                print("시트 이름이 잘못되었습니다.")
        print(f"시트를 읽어왔습니다. 시트 이름 :{sheet_name}")
        sheet = wb[sheet_name]
        print(f"시트의 데이터를 읽어옵니다.")
        print("읽어들일 고객 정보의 시작 행을 입력해주세요. (예 : 2)")
        while True:
            try:
                customer_start_row = int(input())
                if Validator.validate_is_go_back_menu(customer_start_row):
                    return
                else:
                    break
            except ValueError:
                print("숫자만 입력해주세요.")
        print("읽어들일 고객 아이디의 열을 입력해주세요. (예 : A)")
        while True:
            customer_id_col = input()
            if Validator.validate_is_go_back_menu(customer_id_col):
                return
            if Validator.validate_cell(sheet, customer_start_row, customer_id_col):
                Validator.add_to_col_list(customer_id_col)
                break
            else:
                print("잘못된 입력입니다.")
        print("읽어들일 고객 이름의 열을 입력해주세요. (예 : B)")
        while True:
            customer_name_col = input()
            if Validator.validate_is_go_back_menu(customer_name_col):
                return
            if Validator.validate_cell(sheet, customer_start_row, customer_name_col):
                Validator.add_to_col_list(customer_name_col)
                break
            else:
                print("잘못된 입력입니다.")
        print("읽어들일 고객 이용 금액의 열을 입력해주세요. (예 : C)")
        while True:
            customer_price_col = input()
            if Validator.validate_is_go_back_menu(customer_price_col):
                return
            if Validator.validate_cell(sheet, customer_start_row, customer_price_col):
                Validator.add_to_col_list(customer_price_col)
                break
            else:
                print("잘못된 입력입니다.")
        print("읽어들일 고객 이용 시간의 열을 입력해주세요. (예 : D)")
        while True:
            customer_time_col = input()
            if Validator.validate_is_go_back_menu(customer_time_col):
                return
            if Validator.validate_cell(sheet, customer_start_row, customer_time_col):
                Validator.add_to_col_list(customer_time_col)
                break
            else:
                print("잘못된 입력입니다.")
        Validator.clear_row_list()
        print("검증에 실패한 데이터는 입력되지 않습니다.")
        print("고객 정보를 입력합니다.")
        print(f"고객 아이디 열 : {customer_id_col}, 규칙: {Validator.regex_explain_id}")
        print(f"고객 이름 열 : {customer_name_col}, 규칙: {Validator.regex_explain_name}")
        print(f"고객 이용 금액 열 : {customer_price_col}, 규칙: 0 이상의 정수, 0일 시 입력되지 않음")
        print(f"고객 이용 시간 열 : {customer_time_col}, 규칙: 0 이상의 정수, 0일 시 입력되지 않음")
        print(f"고객 정보 시작 행 : {customer_start_row}")
        print("정보를 입력하시겠습니까? (y/n)")
        while True:
            yes_or_no = input()
            if Validator.validate_is_yes_or_no(yes_or_no):
                break
            else:
                print("잘못된 입력입니다.")
        if yes_or_no == "n":
            print("돌아갑니다.")
            return
        print("정보를 입력합니다.")
        self.add_customer_by_xl_work(sheet, customer_id_col, customer_name_col, customer_price_col, customer_time_col,
                                     customer_start_row)
        print("모든 정보를 입력했습니다.")

    def add_customer_by_xl_work(self, sheet, customer_id_col, customer_name_col, customer_price_col, customer_time_col,
                                customer_start_row):
        idx_num = 1
        self.__prev_add_list.clear()
        for i in range(customer_start_row, sheet.max_row + 1):
            try:
                customer_id = sheet[f"{customer_id_col}{i}"].value
                customer_name = sheet[f"{customer_name_col}{i}"].value
                customer_price = int(sheet[f"{customer_price_col}{i}"].value)
                customer_time = int(sheet[f"{customer_time_col}{i}"].value)
                if Validator.validate_customer_id(customer_id) and Validator.validate_customer_name(customer_name) and self.__customer_service.is_exist_by_customer_id(customer_id) is False:
                    if customer_price is None:
                        customer_price = 0
                    if customer_time is None:
                        customer_time = 0
                    customer = Customer.of(customer_id, customer_name, customer_price, customer_time)
                    self.__customer_service.add_customer(customer)
                    self.__prev_add_list.append(customer)
                    print(customer)
                    print(f"{idx_num}번째 고객 정보를 입력했습니다.")
                else:
                    print(f"{idx_num}번째 고객 정보를 입력하지 못했습니다.")
                idx_num += 1
            except ValueError:
                print(f"{idx_num}번째 고객 정보를 입력하지 못했습니다.")
                idx_num += 1

    def print_prev_add_list(self):
        for customer in self.__prev_add_list:
            print(customer)

    def add_customer_recall(self, num: int):
        for i in range(num):
            print(f"{i + 1}번째 고객 정보를 입력하세요.")
            print("이전으로 돌아가기를 원할 시 0을 입력해주세요.")
            while True:
                customer_id = input("고객 ID를 입력해주세요. (알파벳 / 숫자 / _ 가능, 5~12자, 첫글자는 영문자 혹은 숫자, 중복 금지)")
                if Validator.validate_is_go_back_menu(customer_id):
                    return
                if self.__customer_service.is_exist_by_customer_id(customer_id):
                    print(Validator.customer_id_already_exist)
                elif Validator.validate_customer_id(customer_id):
                    break
                else:
                    print(Validator.customer_id_not_valid)
            while True:
                customer_name = input("고객 이름을 입력해주세요. (영문자, 3자 이상)")
                if Validator.validate_is_go_back_menu(customer_name):
                    return
                if Validator.validate_customer_name(customer_name):
                    break
                else:
                    print(Validator.customer_name_not_valid)
            while True:
                try:
                    customer_spent_money = int(input("고객 이용 금액을 만원 단위로 1이상 입력 해주세요. (숫자)"))
                    if Validator.validate_is_go_back_menu(customer_spent_money):
                        return
                    if customer_spent_money < 0:
                        print("0 이상의 숫자를 입력해주세요.")
                    else:
                        break
                except ValueError:
                    print("숫자만 입력해주세요.")
            while True:
                try:
                    customer_spent_hour = int(input("고객 이용 시간을 시간 단위로 1이상 입력 해주세요. (숫자)"))
                    if Validator.validate_is_go_back_menu(customer_spent_hour):
                        return
                    if customer_spent_hour < 0:
                        print("0 이상의 숫자를 입력해주세요.")
                    else:
                        break
                except ValueError:
                    print("숫자만 입력해주세요.")
            print(
                f"고객 ID: {customer_id}, 고객 이름: {customer_name}, 고객 이용 금액: {customer_spent_money}, 고객 이용 시간: {customer_spent_hour}")
            print("이 정보를 추가하시겠습니까? (y/n) 이전 메뉴로 돌아가시려면 0을 입력해주세요.")
            while True:
                reply = input()
                if Validator.validate_is_go_back_menu(reply):
                    return
                if Validator.validate_is_yes_or_no(reply):
                    break
                else:
                    print(Validator.yes_or_no_not_valid)
            if reply == 'y':
                self.__customer_service.add_customer(
                    Customer.of(customer_id, customer_name, customer_spent_money, customer_spent_hour))
                print(f"{i + 1}번째 고객 정보가 추가되었습니다.")
            else:
                print("고객 정보 추가를 취소합니다.")
        print("모든 고객 정보가 추가되었습니다.")

    def print_customers(self):
        if self.__customer_service.get_size() == 0:
            print("고객 명단이 비어있습니다.")
            return
        print("고객 명단을 출력합니다.")
        customers = self.__customer_service.get_customers()
        for customer in customers:
            print(customer)

    def delete_customer_select(self):
        print("1. 고객 명단 직접 삭제")
        print("2. 직전에 엑셀로 입력된 고객 명단 삭제")
        print("3. 이전 메뉴로 돌아가기")
        while True:
            try:
                num = int(input())
                if Validator.validate_is_go_back_menu(num):
                    return
                if num == 1:
                    self.delete_customer()
                    break
                elif num == 2:
                    self.delete_prev_add_customer_xl_select()
                    break
                else:
                    print("잘못된 입력입니다.")
            except ValueError:
                print("숫자만 입력해주세요.")

    def delete_customer(self):
        if self.__customer_service.get_size() == 0:
            print("고객 명단이 비어있습니다.")
            return
        print("고객 명단을 삭제합니다.")
        print("삭제를 원하는 인원 수를 입력해주세요. 돌아가기를 원할 시 0을 입력해주세요.")
        try:
            num = int(input())
            if Validator.validate_is_go_back_menu(num):
                return
        except ValueError:
            print("숫자만 입력해주세요. 돌아갑니다.")
            return
        self.delete_customer_recall(num)

    def delete_customer_recall(self, num):
        for i in range(num):
            print("삭제를 원하는 명단의 고유 번호를 입력해주세요. 돌아가기를 원할 시 0을 입력해주세요.")
            try:
                serial_id = int(input())
                if Validator.validate_is_go_back_menu(serial_id):
                    return
            except ValueError:
                print("숫자만 입력해주세요. 돌아갑니다.")
                return
            if self.__customer_service.is_exist_by_customer_serial_id(serial_id):
                print(self.__customer_service.get_customer_by_customer_serial_id(serial_id))
                print("이 정보를 삭제하시겠습니까? (y/n) 이전 메뉴로 돌아가시려면 0을 입력해주세요.")
                while True:
                    reply = input()
                    if Validator.validate_is_go_back_menu(reply):
                        return
                    if Validator.validate_is_yes_or_no(reply):
                        if reply == 'y':
                            self.__customer_service.delete_customer_by_customer_serial_id(serial_id)
                            print(f"{serial_id}번째 고객 정보가 삭제되었습니다.")
                            break
                        else:
                            print("고객 정보 삭제를 취소합니다.")
                            break
                    else:
                        print(Validator.yes_or_no_not_valid)
            else:
                print("해당 고유 번호의 고객 정보가 존재하지 않습니다.")
        print("모든 고객 정보가 삭제되었습니다.")

    def delete_prev_add_customer_xl_select(self):
        if self.__prev_add_list.__len__() == 0:
            print("이전에 입력된 고객 정보가 없습니다.")
            return
        print("이전에 엑셀로 입력한 고객 정보를 삭제합니다.")
        print("이전에 입력된 고객 정보입니다.")
        self.print_prev_add_list()
        print("정말로 삭제하시겠습니까? (y/n)")
        while True:
            yes_or_no = input()
            if Validator.validate_is_yes_or_no(yes_or_no):
                break
            else:
                print("잘못된 입력입니다.")
        if yes_or_no == "n":
            print("돌아갑니다.")
            return
        if yes_or_no == "y":
            self.delete_prev_add_customer()
            return

    def delete_prev_add_customer(self):
        for customer in self.__prev_add_list:
            self.__customer_service.delete_customer_by_customer_serial_id(customer.customer_serial_id)
        print("모든 정보를 삭제했습니다.")
        self.__prev_add_list.clear()


    def update_customer(self):
        if self.__customer_service.get_size() == 0:
            print("고객 명단이 비어있습니다.")
            return
        print("고객 명단을 수정합니다.")
        print(f"수정을 원하는 명단의 고유 번호를 숫자로 입력해주세요. (1~{self.__customer_service.get_size()}번)")
        try:
            serial_id = int(input())
            if self.__customer_service.is_exist_by_customer_serial_id(serial_id):
                customer = self.__customer_service.get_customer_by_customer_serial_id(serial_id)
                self.select_update_menu(customer)
            else:
                print("해당 고유 번호의 고객 정보가 존재하지 않습니다.")

        except ValueError:
            print("숫자만 입력해주세요. 돌아갑니다.")
            return

    def update_customer_id(self, customer: Customer):
        print(f"현재 고객의 아이디는 : {customer.customer_id} 입니다.")
        print(f"변경을 원하는 아이디를 입력해주세요. ({Validator.regex_explain_id})")
        print("돌아가시려면 0을 입력해주세요.")
        while True:
            input_customer_id = input()
            if Validator.validate_is_go_back_menu(input_customer_id):
                return
            if Validator.validate_customer_id(input_customer_id):
                if self.__customer_service.is_exist_by_customer_id(input_customer_id):
                    print(Validator.customer_id_already_exist)
                else:
                    customer.customer_id = input_customer_id
                    self.__customer_service.edit_customer_by_customer_serial_id(customer.customer_serial_id, customer)
                    print(f"고객의 아이디가 {customer.customer_id}로 변경되었습니다.")
                    break
            else:
                print(Validator.customer_id_not_valid)

    def update_customer_name(self, customer: Customer):
        print(f"현재 입력된 고객의 이름은 : {customer.customer_name} 입니다.")
        print(f"변경을 원하는 이름을 입력해주세요. ({Validator.regex_explain_name})")
        print("돌아가시려면 0을 입력해주세요.")
        while True:
            input_customer_name = input()
            if Validator.validate_is_go_back_menu(input_customer_name):
                return
            if Validator.validate_customer_name(input_customer_name):
                customer.customer_name = input_customer_name
                self.__customer_service.edit_customer_by_customer_serial_id(customer.customer_serial_id, customer)
                print(f"고객의 이름이 {customer.customer_name} 으로 변경되었습니다.")
                break
            else:
                print(Validator.customer_name_not_valid)

    def update_customer_spent_money(self, customer: Customer):
        print(f"현재 입력된 고객의 이용 금액은 : {customer.customer_spent_money} 입니다.")
        print(f"변경을 원하는 이용 금액을 만원 단위로 1 이상 입력해주세요.")
        print("돌아가시려면 0을 입력해주세요.")
        while True:
            input_customer_spent_money = input()
            if Validator.validate_is_go_back_menu(input_customer_spent_money):
                return
            else:
                customer.customer_spent_money = input_customer_spent_money
                self.__customer_service.edit_customer_by_customer_serial_id(customer.customer_serial_id, customer)
                print(f"고객의 이용 금액이 {customer.customer_spent_money} 으로 변경되었습니다.")
                break

    def update_customer_spent_hour(self, customer: Customer):
        print(f"현재 입력된 고객의 이용 시간은 : {customer.customer_spent_hour} 입니다.")
        print(f"변경을 원하는 이용 시간을 시간 단위로 1 이상 입력해주세요.")
        print("돌아가시려면 0을 입력해주세요.")
        while True:
            input_customer_spent_hour = input()
            if Validator.validate_is_go_back_menu(input_customer_spent_hour):
                return
            else:
                customer.customer_spent_hour = input_customer_spent_hour
                self.__customer_service.edit_customer_by_customer_serial_id(customer.customer_serial_id, customer)
                print(f"고객의 이용 시간이 {customer.customer_spent_hour} 으로 변경되었습니다.")
                break

    def select_update_menu(self, customer: Customer):
        while True:
            try:
                print(customer)
                print("수정을 원하는 항목을 선택해주세요.")
                print("1. 고객 ID, 2. 고객 이름, 3. 고객 이용 금액, 4. 고객 이용 시간")
                print("돌아가시려면 0을 입력해주세요.")
                menu = int(input())
                if Validator.validate_is_go_back_menu(menu):
                    return
                if menu == 1:
                    self.update_customer_id(customer)
                    break
                if menu == 2:
                    self.update_customer_name(customer)
                    break
            except ValueError:
                print("숫자만 입력해주세요.")
                return
