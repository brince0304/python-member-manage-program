import re
import string

import openpyxl.utils.exceptions
from openpyxl.reader.excel import load_workbook


class Validator:
    regex_customer_name = re.compile("^[a-zA-Z]{3,}$")
    regex_customer_id = re.compile("/^[a-zA-Z][a-zA-Z0-9_]{4,11}$/")
    customer_name_not_valid = "고객 이름이 유효하지 않습니다."
    customer_id_not_valid = "고객 아이디가 유효하지 않습니다."
    go_back_menu = "이전 메뉴로 돌아갑니다."
    yes_or_no_not_valid = "y 혹은 n만 입력해주세요."
    customer_id_already_exist = "이미 존재하는 고객 아이디입니다."
    regex_explain_id = "알파벳 / 숫자 / _ 가능, 5~12자, 첫글자는 영문자 혹은 숫자, 중복 금지"
    regex_explain_name = "영문자, 3자 이상"
    col_list = list()

    @staticmethod
    def validate_customer_id(customer_id):
        if Validator.regex_customer_id.match(customer_id) is None:
            return False
        else:
            return True

    @staticmethod
    def validate_customer_name(customer_name):
        if Validator.regex_customer_name.match(customer_name) is None:
            return False
        else:
            return True

    @staticmethod
    def validate_is_go_back_menu(input_menu):
        if input_menu == "0" or input_menu == 0:
            print(Validator.go_back_menu)
            Validator.clear_row_list()
            return True
        else:
            return False

    @staticmethod
    def validate_is_yes_or_no(input_menu):
        ignore_case = input_menu.lower()
        if ignore_case == "y" or ignore_case == "n":
            return True
        else:
            return False

    @staticmethod
    def validate_file_path(file_path):
        try:
            wb = load_workbook(file_path)
            if wb is None:
                return False
            else:
                return True
        except openpyxl.utils.exceptions.InvalidFileException:
            return False

    @staticmethod
    def validate_sheet_name(wb, sheet_name):
        if sheet_name in wb.sheetnames:
            return True
        else:
            return False

    @staticmethod
    def validate_cell(sheet, row, col):
        try:
            if col in Validator.col_list:
                return False
            if sheet.cell(row=row, column=Validator.validate_convert_col_to_int(col)).value is None:
                return False
            else:
                return True
        except IndexError:
            return False
        except TypeError:
            return False

    @staticmethod
    def clear_row_list():
        Validator.col_list.clear()

    @staticmethod
    def validate_col(sheet, customer_start_row, customer_id_col, customer_name_col, customer_spent_money_col,
                     customer_spent_hour_col):
        if sheet.cell(row=customer_start_row, column=customer_id_col).value is None or \
                sheet.cell(row=customer_start_row, column=customer_name_col).value is None or \
                sheet.cell(row=customer_start_row, column=customer_spent_money_col).value is None or \
                sheet.cell(row=customer_start_row, column=customer_spent_hour_col).value is None:
            return False
        else:
            return True

    @staticmethod
    def validate_convert_col_to_int(col: str):
        try:
            return ord(col) - 64 if col.isupper() else ord(col) - 96
        except TypeError:
            return col

    @staticmethod
    def validate_convert_col_to_char(col: str):
        return chr(col + 64) if col <= 26 else chr(col + 70)

    @staticmethod
    def add_to_col_list(row):
        Validator.col_list.append(row)
